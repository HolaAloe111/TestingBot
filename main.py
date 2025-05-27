import asyncio
import logging
import random
import os
import time
import config  # Assuming this file contains BOT_TOKEN
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from aiogram import Bot, Dispatcher, F
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    Message,
    CallbackQuery,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    FSInputFile,
)
from aiogram.filters import Command, CommandStart, StateFilter
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.context import FSMContext
from aiogram.exceptions import TelegramBadRequest

# --- Globals & Constants ---
bot = Bot(
    config.BOT_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML)
)
dp = Dispatcher(storage=MemoryStorage())

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
)
logger = logging.getLogger(__name__)

EXCEL_FILENAME = "persistent_user_data.xlsx"

BASE_HEADERS = ["Telegram ID", "Unique ID", "Name", "Age"]
CORSI_HEADERS = [
    "Corsi - Max Correct Sequence Length",
    "Corsi - Avg Time Per Element (s)",
    "Corsi - Sequence Times Detail",
    "Corsi - Interrupted",
]
STROOP_HEADERS = [
    "Stroop Part1 Time (s)", "Stroop Part1 Errors",
    "Stroop Part2 Time (s)", "Stroop Part2 Errors",
    "Stroop Part3 Time (s)", "Stroop Part3 Errors",
    "Stroop - Interrupted",
]
ALL_EXPECTED_HEADERS = BASE_HEADERS + CORSI_HEADERS + STROOP_HEADERS

IKB = InlineKeyboardButton

ACTION_SELECTION_KEYBOARD_NEW = InlineKeyboardMarkup(
    inline_keyboard=[
        [IKB(text="Пройти батарею тестов", callback_data="run_test_battery")],
        [IKB(text="Выбрать отдельный тест", callback_data="select_specific_test")],
    ]
)

ACTION_SELECTION_KEYBOARD_RETURNING = InlineKeyboardMarkup(
    inline_keyboard=[
        [IKB(text="Пройти батарею тестов заново", callback_data="run_test_battery")],
        [IKB(text="Выбрать отдельный тест заново", callback_data="select_specific_test")],
        [IKB(text="Выйти (сбросить профиль)", callback_data="logout_profile")]
    ]
)


# --- FSM States ---
class UserData(StatesGroup):
    waiting_for_first_time_response = State()
    waiting_for_name = State()
    waiting_for_age = State()
    waiting_for_unique_id = State()
    waiting_for_test_overwrite_confirmation = State()


class CorsiTestStates(StatesGroup):
    showing_sequence = State()
    waiting_for_user_sequence = State()


class StroopTestStates(StatesGroup):
    part1_display = State()
    part1_response = State()
    part2_display = State()
    part2_response = State()
    part3_display = State()
    part3_response = State()


# --- Helper Functions ---
def initialize_excel_file():
    if not os.path.exists(EXCEL_FILENAME):
        wb = Workbook()
        ws = wb.active
        ws.append(ALL_EXPECTED_HEADERS)
        wb.save(EXCEL_FILENAME)
        logger.info(f"'{EXCEL_FILENAME}' created with all headers.")
    else:
        try:
            wb = load_workbook(EXCEL_FILENAME)
            ws = wb.active
            if ws.max_row == 0:
                ws.append(ALL_EXPECTED_HEADERS)
                logger.info(f"Appended all headers to empty sheet in '{EXCEL_FILENAME}'.")
            else:
                current_headers = [cell.value for cell in ws[1]]
                new_headers_to_add = [h for h in ALL_EXPECTED_HEADERS if h not in current_headers]
                if new_headers_to_add:
                    header_col_start_index = len(current_headers) + 1
                    for i, header in enumerate(new_headers_to_add):
                        ws.cell(row=1, column=header_col_start_index + i).value = header
                    logger.info(f"Added missing headers to '{EXCEL_FILENAME}': {new_headers_to_add}")
            wb.save(EXCEL_FILENAME)
            logger.info(f"'{EXCEL_FILENAME}' checked/updated for headers.")
        except (InvalidFileException, Exception) as e:
            logger.error(
                f"Error initializing/updating Excel file '{EXCEL_FILENAME}': {e}. Manual check might be needed.")


async def get_active_profile_from_fsm(state: FSMContext) -> dict | None:
    data = await state.get_data()
    if data.get("active_unique_id"):
        return {
            "unique_id": data.get("active_unique_id"),
            "name": data.get("active_name"),
            "age": data.get("active_age"),
            "telegram_id": data.get("active_telegram_id"),
        }
    return None


async def send_main_action_menu(
        trigger_event_or_message: [Message, CallbackQuery],
        keyboard_markup: InlineKeyboardMarkup,
        text: str = "Выберите дальнейшее действие:",
        state: FSMContext = None
):
    chat_id = None
    if isinstance(trigger_event_or_message, Message):
        chat_id = trigger_event_or_message.chat.id
    elif isinstance(trigger_event_or_message, CallbackQuery):
        chat_id = trigger_event_or_message.message.chat.id
        try:
            await trigger_event_or_message.message.edit_reply_markup(reply_markup=None)
        except TelegramBadRequest:
            pass

    if chat_id:
        try:
            await bot.send_message(chat_id, text, reply_markup=keyboard_markup)
        except Exception as e:
            logger.error(f"Error in send_main_action_menu for chat {chat_id}: {e}")


# --- Corsi Test Specific Logic ---
async def cleanup_corsi_messages(state: FSMContext, bot_instance: Bot, final_text: str = None):
    data = await state.get_data()
    chat_id = data.get('corsi_chat_id')
    if not chat_id:
        logger.warning("No 'corsi_chat_id' in FSM for Corsi cleanup, or it was already cleared.")
        return

    msg_ids_to_delete = []
    if data.get('corsi_status_message_id'):
        msg_ids_to_delete.append(data.get('corsi_status_message_id'))
    if data.get('corsi_feedback_message_id'):
        msg_ids_to_delete.append(data.get('corsi_feedback_message_id'))

    for msg_id in msg_ids_to_delete:
        if msg_id:
            try:
                await bot_instance.delete_message(chat_id=chat_id, message_id=msg_id)
                logger.info(f"Deleted Corsi message {msg_id} in chat {chat_id}")
            except TelegramBadRequest:
                logger.warning(f"Failed to delete Corsi message {msg_id} in chat {chat_id} (likely already deleted).")
                pass

    grid_message_id = data.get('corsi_grid_message_id')
    if grid_message_id:
        try:
            text_to_set = final_text if final_text else "Тест Корси завершен или отменен."
            await bot_instance.edit_message_text(
                text=text_to_set, chat_id=chat_id, message_id=grid_message_id, reply_markup=None
            )
            logger.info(f"Edited Corsi grid message {grid_message_id} in chat {chat_id}")
        except TelegramBadRequest:
            logger.warning(
                f"Failed to edit Corsi grid message {grid_message_id} in chat {chat_id} (likely already deleted or no change).")
            pass

    corsi_operational_keys_for_fsm_cleanup = [
        'corsi_chat_id', 'corsi_status_message_id', 'corsi_feedback_message_id', 'corsi_grid_message_id',
        'current_sequence_length', 'error_count', 'sequence_times', 'correct_sequence',
        'user_input_sequence', 'sequence_start_time'
    ]
    current_fsm_data = await state.get_data()
    data_after_corsi_message_cleanup = {k: v for k, v in current_fsm_data.items() if
                                        k not in corsi_operational_keys_for_fsm_cleanup}
    await state.set_data(data_after_corsi_message_cleanup)
    logger.info(f"Cleared Corsi operational FSM keys for chat {chat_id}")


async def show_corsi_sequence(trigger_message: Message, state: FSMContext):
    data = await state.get_data()
    if await state.get_state() != CorsiTestStates.showing_sequence.state:  # Crucial state check
        logger.info(f"show_corsi_sequence called but state is {await state.get_state()}. Aborting.")
        return

    current_sequence_length = data.get('current_sequence_length', 2)
    corsi_chat_id = data.get('corsi_chat_id')
    if not corsi_chat_id:
        logger.error("corsi_chat_id not in FSM at start of show_corsi_sequence. Aborting.")
        await state.clear()
        await trigger_message.answer("Произошла ошибка с тестом Корси. Пожалуйста, начните заново с /start.")
        return

    grid_message_id_from_state = data.get('corsi_grid_message_id')

    restart_button_row = [IKB(text="🔄", callback_data="corsi_stop_this_attempt")]
    button_indices = list(range(9))
    random.shuffle(button_indices)
    correct_sequence = button_indices[:current_sequence_length]
    await state.update_data(correct_sequence=correct_sequence, user_input_sequence=[])

    base_buttons = [IKB(text="🟪", callback_data=f"corsi_button_{i}") for i in range(9)]
    base_keyboard_grid_rows = [base_buttons[i:i + 3] for i in range(0, 9, 3)]
    keyboard_for_base_markup_with_restart = [row[:] for row in base_keyboard_grid_rows]
    keyboard_for_base_markup_with_restart.append(restart_button_row)
    base_markup_with_restart = InlineKeyboardMarkup(inline_keyboard=keyboard_for_base_markup_with_restart)

    if grid_message_id_from_state:
        try:
            await bot.edit_message_text(
                chat_id=corsi_chat_id, message_id=grid_message_id_from_state,
                text="Тест Корси", reply_markup=base_markup_with_restart
            )
        except TelegramBadRequest:
            grid_msg_obj = await bot.send_message(corsi_chat_id, "Тест Корси", reply_markup=base_markup_with_restart)
            grid_message_id_from_state = grid_msg_obj.message_id
    else:
        grid_msg_obj = await bot.send_message(corsi_chat_id, "Тест Корси", reply_markup=base_markup_with_restart)
        grid_message_id_from_state = grid_msg_obj.message_id
    await state.update_data(corsi_grid_message_id=grid_message_id_from_state)

    corsi_status_message_id = data.get('corsi_status_message_id')
    status_text_queue = ["Приготовьтесь..."] + [f"{i}..." for i in range(3, 0, -1)] + ["Запоминайте..."]

    for i, text in enumerate(status_text_queue):
        if await state.get_state() != CorsiTestStates.showing_sequence.state:
            logger.info(f"Corsi state changed during status message display ({text}); aborting loop.")
            return

        if not corsi_status_message_id:
            try:
                status_obj = await bot.send_message(corsi_chat_id, text)
                corsi_status_message_id = status_obj.message_id
                await state.update_data(corsi_status_message_id=corsi_status_message_id)
            except Exception as e:
                logger.error(f"Error sending initial Corsi status message: {e}")
                return
        else:
            try:
                await bot.edit_message_text(text=text, chat_id=corsi_chat_id, message_id=corsi_status_message_id)
            except TelegramBadRequest:
                logger.warning(
                    f"Corsi status message {corsi_status_message_id} not found for edit ({text}), breaking loop.")
                return
            except Exception as e:
                logger.error(f"Error editing Corsi status message ({text}): {e}")
                return

        if i < len(status_text_queue) - 1:
            await asyncio.sleep(1)
        else:
            await asyncio.sleep(0.5)

    if await state.get_state() != CorsiTestStates.showing_sequence.state:
        logger.info("Corsi state changed before flashing sequence; aborting.")
        return

    for button_index in correct_sequence:
        if await state.get_state() != CorsiTestStates.showing_sequence.state:
            logger.info(f"Corsi state changed during flash sequence; aborting loop.")
            return

        flashed_rows = [
            [IKB(text="🟨" if r * 3 + c == button_index else "🟪", callback_data=f"corsi_button_{r * 3 + c}") for c in
             range(3)] for r in range(3)]
        flashed_rows.append(list(restart_button_row))
        flashed_markup = InlineKeyboardMarkup(inline_keyboard=flashed_rows)
        try:
            await bot.edit_message_reply_markup(chat_id=corsi_chat_id, message_id=grid_message_id_from_state,
                                                reply_markup=flashed_markup)
            await asyncio.sleep(0.5)
            await bot.edit_message_reply_markup(chat_id=corsi_chat_id, message_id=grid_message_id_from_state,
                                                reply_markup=base_markup_with_restart)
            await asyncio.sleep(0.2)
        except TelegramBadRequest:
            logger.warning(f"Corsi grid message {grid_message_id_from_state} not found for flash edit, breaking loop.")
            return
        except Exception as e:
            logger.error(f"Error during Corsi flash sequence: {e}")
            return

    if await state.get_state() != CorsiTestStates.showing_sequence.state:
        logger.info("Corsi state changed before prompting user input; aborting.")
        return

    try:
        current_data_for_final_prompt = await state.get_data()
        status_msg_id_for_final_prompt = current_data_for_final_prompt.get('corsi_status_message_id')
        if status_msg_id_for_final_prompt:
            await bot.edit_message_text(text="Повторите последовательность:", chat_id=corsi_chat_id,
                                        message_id=status_msg_id_for_final_prompt)
        else:
            logger.warning("Corsi status message ID was None before final prompt. Re-sending prompt message.")
            status_obj = await bot.send_message(corsi_chat_id, "Повторите последовательность:")
            await state.update_data(corsi_status_message_id=status_obj.message_id)


    except TelegramBadRequest:
        logger.warning("Corsi status message not found for final prompt 'Повторите последовательность:'.")
        return
    except Exception as e:
        logger.error(f"Error editing Corsi status message for final prompt: {e}")
        return

    await state.update_data(sequence_start_time=time.time())
    await state.set_state(CorsiTestStates.waiting_for_user_sequence)


async def handle_corsi_button_press(callback: CallbackQuery, state: FSMContext):
    if await state.get_state() != CorsiTestStates.waiting_for_user_sequence.state:
        await callback.answer("Тест был прерван или завершен.", show_alert=True)
        logger.warning("handle_corsi_button_press called but state is not waiting_for_user_sequence.")
        return

    await callback.answer()
    button_index = int(callback.data.split("_")[-1])
    data = await state.get_data()
    user_input_sequence = data.get('user_input_sequence', []) + [button_index]

    corsi_grid_message_id = data.get('corsi_grid_message_id')
    corsi_chat_id = data.get('corsi_chat_id')
    if not corsi_grid_message_id or not corsi_chat_id:
        logger.error("Corsi grid/chat ID missing in handle_corsi_button_press.")
        await callback.message.answer(
            "Ошибка: не удалось обработать ваш ввод. Пожалуйста, попробуйте /start и начните тест заново.")
        await state.clear()
        return

    new_rows = [
        [IKB(text="🟨" if r * 3 + c in user_input_sequence else "🟪", callback_data=f"corsi_button_{r * 3 + c}") for c in
         range(3)] for r in range(3)]
    new_rows.append([IKB(text="🔄", callback_data="corsi_stop_this_attempt")])
    try:
        await bot.edit_message_reply_markup(
            chat_id=corsi_chat_id, message_id=corsi_grid_message_id,
            reply_markup=InlineKeyboardMarkup(inline_keyboard=new_rows)
        )
    except TelegramBadRequest as e:
        logger.error(f"Error editing markup on Corsi button press: {e}")
        return

    await state.update_data(user_input_sequence=user_input_sequence)
    if await state.get_state() == CorsiTestStates.waiting_for_user_sequence.state:
        if len(user_input_sequence) == len(data.get('correct_sequence', [])):
            await evaluate_user_sequence(callback.message, state)


async def on_corsi_restart_current_test(callback: CallbackQuery, state: FSMContext):
    await callback.answer(text='Тест Корси будет прерван.', show_alert=False)
    await stop_test_command_handler(callback.message, state, called_from_test_button=True)


async def evaluate_user_sequence(message_context: Message, state: FSMContext):
    if await state.get_state() != CorsiTestStates.waiting_for_user_sequence.state:
        logger.warning("evaluate_user_sequence called but state is not waiting_for_user_sequence.")
        return

    data = await state.get_data()
    chat_id = data.get('corsi_chat_id', message_context.chat.id)
    user_seq = data.get('user_input_sequence', [])
    correct_seq = data.get('correct_sequence', [])
    curr_len = data.get('current_sequence_length', 2)
    err_count = data.get('error_count', 0)
    seq_times = data.get('sequence_times', [])
    start_time = data.get('sequence_start_time', 0)
    fb_id = data.get('corsi_feedback_message_id')
    time_taken = time.time() - start_time

    feedback_message_text = ""
    test_continues = True

    if user_seq == correct_seq:
        seq_times.append({'len': curr_len, 'time': time_taken})
        curr_len += 1
        err_count = 0
        feedback_message_text = "<b>Верно!</b>"
        if curr_len > 9:
            test_continues = False
    else:
        err_count += 1
        feedback_message_text = "<b>Ошибка! Попробуйте ещё раз.</b>"
        if err_count >= 2:
            test_continues = False

    if fb_id:
        try:
            await bot.edit_message_text(feedback_message_text, chat_id=chat_id, message_id=fb_id,
                                        parse_mode=ParseMode.HTML)
        except TelegramBadRequest:
            fb_id = None
    if not fb_id:
        try:
            fb_msg_obj = await bot.send_message(chat_id, feedback_message_text, parse_mode=ParseMode.HTML)
            fb_id = fb_msg_obj.message_id
            await state.update_data(corsi_feedback_message_id=fb_id)
        except Exception as e:
            logger.error(f"Error sending feedback message in evaluate_user_sequence: {e}")
            pass

    if user_seq == correct_seq and test_continues and fb_id:
        await asyncio.sleep(0.7)
        try:
            await bot.edit_message_text("Верно!", chat_id=chat_id, message_id=fb_id, parse_mode=None)
        except TelegramBadRequest:
            pass

    if await state.get_state() != CorsiTestStates.waiting_for_user_sequence.state:
        logger.info(
            "State changed during evaluate_user_sequence before deciding next step. Aborting further action here.")
        return

    if test_continues:
        await state.set_state(CorsiTestStates.showing_sequence)  # <<<--- FIX: SET STATE HERE
        await state.update_data(
            current_sequence_length=curr_len, error_count=err_count,
            sequence_times=seq_times, user_input_sequence=[]
        )
        await show_corsi_sequence(message_context, state)
    else:
        await save_corsi_results(message_context, state, is_interrupted=False)
        await cleanup_corsi_messages(state, bot, final_text="Тест Корси завершен.")

        fsm_data_after_test_cleanup = await state.get_data()
        main_profile_data_to_keep = {}
        if "active_unique_id" in fsm_data_after_test_cleanup:
            main_profile_data_to_keep["active_unique_id"] = fsm_data_after_test_cleanup.get("active_unique_id")
            main_profile_data_to_keep["active_name"] = fsm_data_after_test_cleanup.get("active_name")
            main_profile_data_to_keep["active_age"] = fsm_data_after_test_cleanup.get("active_age")
            main_profile_data_to_keep["active_telegram_id"] = fsm_data_after_test_cleanup.get("active_telegram_id")

        await state.set_state(None)
        if main_profile_data_to_keep.get("active_unique_id"):
            await state.set_data(main_profile_data_to_keep)
            await send_main_action_menu(message_context, ACTION_SELECTION_KEYBOARD_RETURNING, state=state)
        else:
            await message_context.answer("Тест завершен, но ваш профиль не активен. Пожалуйста, используйте /start.")
            await state.clear()


async def start_corsi_test(trigger_event_or_message: [Message, CallbackQuery], state: FSMContext, profile_data: dict):
    logger.info(f"Starting Corsi Test for UID: {profile_data.get('unique_id')}")

    message_context = trigger_event_or_message.message if isinstance(trigger_event_or_message,
                                                                     CallbackQuery) else trigger_event_or_message
    await state.set_state(CorsiTestStates.showing_sequence)
    await state.update_data(
        unique_id_for_test=profile_data.get('unique_id'),
        profile_name_for_test=profile_data.get('name'),
        profile_age_for_test=profile_data.get('age'),
        profile_telegram_id_for_test=profile_data.get('telegram_id'),
        current_sequence_length=2, error_count=0, sequence_times=[],
        correct_sequence=[], user_input_sequence=[], sequence_start_time=0,
        corsi_grid_message_id=None, corsi_status_message_id=None,
        corsi_chat_id=message_context.chat.id, corsi_feedback_message_id=None,
    )
    await show_corsi_sequence(message_context, state)


async def save_corsi_results(trigger_event_message: Message, state: FSMContext, is_interrupted: bool = False):
    data = await state.get_data()
    unique_id = data.get('unique_id_for_test')
    profile_telegram_id = data.get('profile_telegram_id_for_test')
    profile_name = data.get('profile_name_for_test')
    profile_age = data.get('profile_age_for_test')

    if not unique_id:
        current_state_for_error = await state.get_state()
        logger.error(
            f"CRITICAL: 'unique_id_for_test' not found in FSM state for Corsi save (State: {current_state_for_error}). Interactor: {trigger_event_message.chat.id}. FSM Data: {data}")
        if current_state_for_error is not None:
            await trigger_event_message.answer("Тест Корси: критическая ошибка при сохранении (ID профиля не найден).")
        return

    sequence_times = data.get('sequence_times', [])
    corsi_max_len = max(item['len'] for item in sequence_times) if sequence_times else 0
    corsi_avg_time_per_element = 0.0
    if sequence_times:
        valid_sequences = [item for item in sequence_times if item['len'] > 0]
        if valid_sequences:
            total_avg_time_sum = sum(item['time'] / item['len'] for item in valid_sequences)
            corsi_avg_time_per_element = total_avg_time_sum / len(valid_sequences)
    corsi_detail_string = "; ".join([f"L{item['len']}:{item['time']:.2f}s" for item in sequence_times])
    interruption_status = "Да" if is_interrupted else "Нет"

    try:
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        row_to_update = -1
        for idx, row_cells in enumerate(ws.iter_rows(min_row=2), start=2):
            if row_cells[1].value == unique_id:
                row_to_update = idx
                break

        if row_to_update == -1:
            logger.error(
                f"UID {unique_id} for Corsi results not found in Excel. Appending profile data along with test results.")
            new_row_data = [''] * len(ALL_EXPECTED_HEADERS)
            new_row_data[ALL_EXPECTED_HEADERS.index("Telegram ID")] = profile_telegram_id
            new_row_data[ALL_EXPECTED_HEADERS.index("Unique ID")] = unique_id
            new_row_data[ALL_EXPECTED_HEADERS.index("Name")] = profile_name
            new_row_data[ALL_EXPECTED_HEADERS.index("Age")] = profile_age
            ws.append(new_row_data)
            row_to_update = ws.max_row

        ws.cell(row=row_to_update,
                column=ALL_EXPECTED_HEADERS.index("Corsi - Max Correct Sequence Length") + 1).value = corsi_max_len
        ws.cell(row=row_to_update,
                column=ALL_EXPECTED_HEADERS.index("Corsi - Avg Time Per Element (s)") + 1).value = round(
            corsi_avg_time_per_element, 2)
        ws.cell(row=row_to_update,
                column=ALL_EXPECTED_HEADERS.index("Corsi - Sequence Times Detail") + 1).value = corsi_detail_string
        ws.cell(row=row_to_update,
                column=ALL_EXPECTED_HEADERS.index("Corsi - Interrupted") + 1).value = interruption_status

        wb.save(EXCEL_FILENAME)
        logger.info(f"Corsi results for UID {unique_id} saved/updated. Interrupted: {is_interrupted}")

        current_state_for_summary = await state.get_state()
        if current_state_for_summary is not None:
            summary_text = (
                f"Тест Корси {'<b>ПРЕРВАН</b>' if is_interrupted else '<b>ЗАВЕРШЕН</b>'}!\n"
                f"Максимальная длина последовательности: {corsi_max_len}\n"
                f"Среднее время на элемент: {round(corsi_avg_time_per_element, 2)} сек\n"
                f"Детализация: {corsi_detail_string}"
            )
            if is_interrupted and corsi_max_len == 0 and not sequence_times:
                summary_text = f"Тест Корси <b>ПРЕРВАН</b> досрочно. Результаты не зафиксированы."
            await trigger_event_message.answer(summary_text, parse_mode=ParseMode.HTML)

    except Exception as e:
        logger.error(f"Error saving Corsi results to Excel for UID {unique_id}: {e}")
        current_state_for_error_msg = await state.get_state()
        if current_state_for_error_msg is not None:
            await trigger_event_message.answer("Произошла ошибка при сохранении результатов Теста Корси.")


# --- Stroop Test Skeletons ---
async def start_stroop_test(trigger_event_or_message: [Message, CallbackQuery], state: FSMContext, profile_data: dict):
    logger.info(f"Placeholder: Starting Stroop Test for UID: {profile_data.get('unique_id')}")
    message_context = trigger_event_or_message.message if isinstance(trigger_event_or_message,
                                                                     CallbackQuery) else trigger_event_or_message

    await state.set_state(StroopTestStates.part1_display)
    await state.update_data(
        unique_id_for_test=profile_data.get('unique_id'),
        profile_name_for_test=profile_data.get('name'),
        profile_age_for_test=profile_data.get('age'),
        profile_telegram_id_for_test=profile_data.get('telegram_id'),
        stroop_part1_time_total=None, stroop_part1_errors_total=None,
        stroop_part2_time_total=None, stroop_part2_errors_total=None,
        stroop_part3_time_total=None, stroop_part3_errors_total=None,
        stroop_current_trial=0,
        stroop_stimuli_colors=["Красный", "Синий", "Зеленый", "Желтый", "Черный"],
        stroop_chat_id=message_context.chat.id,
        stroop_main_message_id=None,
    )
    msg = await message_context.answer(
        f"Тест Струпа (Часть 1 - в разработке).\nНажмите кнопку, чтобы 'начать'.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [IKB(text="Начать часть 1 (пример)", callback_data="stroop_p1_next")]
        ]))
    await state.update_data(stroop_main_message_id=msg.message_id)


async def handle_stroop_part1_response(callback: CallbackQuery, state: FSMContext):
    logger.info("Placeholder: Handling Stroop Part 1 response.")
    await callback.answer("Часть 1 Струпа - ответ получен (пример).")

    await save_stroop_results(callback.message, state, is_interrupted=False)
    await cleanup_stroop_ui(state, bot, final_text="Тест Струпа (пример) завершен.")

    fsm_data_after_test_cleanup = await state.get_data()
    main_profile_data_to_keep = {}
    if "active_unique_id" in fsm_data_after_test_cleanup:
        main_profile_data_to_keep["active_unique_id"] = fsm_data_after_test_cleanup.get("active_unique_id")
        main_profile_data_to_keep["active_name"] = fsm_data_after_test_cleanup.get("active_name")
        main_profile_data_to_keep["active_age"] = fsm_data_after_test_cleanup.get("active_age")
        main_profile_data_to_keep["active_telegram_id"] = fsm_data_after_test_cleanup.get("active_telegram_id")

    await state.set_state(None)
    if main_profile_data_to_keep.get("active_unique_id"):
        await state.set_data(main_profile_data_to_keep)
        await send_main_action_menu(callback.message, ACTION_SELECTION_KEYBOARD_RETURNING, state=state)
    else:
        await callback.message.answer("Тест завершен, но профиль не найден. /start")
        await state.clear()


async def save_stroop_results(trigger_event_message: Message, state: FSMContext, is_interrupted: bool = False):
    logger.info(f"Saving Stroop Test results. Interrupted: {is_interrupted}")
    data = await state.get_data()
    unique_id = data.get('unique_id_for_test')
    profile_telegram_id = data.get('profile_telegram_id_for_test')
    profile_name = data.get('profile_name_for_test')
    profile_age = data.get('profile_age_for_test')

    if not unique_id:
        current_state_for_error = await state.get_state()
        logger.error(
            f"CRITICAL: 'unique_id_for_test' not found for Stroop save (State: {current_state_for_error}). Interactor: {trigger_event_message.chat.id}. FSM Data: {data}")
        if current_state_for_error is not None:
            await trigger_event_message.answer("Тест Струпа: критическая ошибка при сохранении (ID профиля не найден).")
        return

    p1_time = data.get("stroop_part1_time_total")
    p1_errors = data.get("stroop_part1_errors_total")
    p2_time = data.get("stroop_part2_time_total")
    p2_errors = data.get("stroop_part2_errors_total")
    p3_time = data.get("stroop_part3_time_total")
    p3_errors = data.get("stroop_part3_errors_total")
    interruption_status_stroop = "Да" if is_interrupted else "Нет"

    try:
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        row_to_update = -1
        for idx, row_cells in enumerate(ws.iter_rows(min_row=2), start=2):
            if row_cells[1].value == unique_id:
                row_to_update = idx
                break
        if row_to_update == -1:
            logger.error(
                f"UID {unique_id} for Stroop results not found in Excel. Appending profile data along with test results.")
            new_row_data = [''] * len(ALL_EXPECTED_HEADERS)
            new_row_data[
                ALL_EXPECTED_HEADERS.index("Telegram ID")] = profile_telegram_id if profile_telegram_id else data.get(
                'active_telegram_id', 'N/A_ExcelError')
            new_row_data[ALL_EXPECTED_HEADERS.index("Unique ID")] = unique_id
            new_row_data[ALL_EXPECTED_HEADERS.index("Name")] = profile_name if profile_name else data.get('active_name',
                                                                                                          'N/A_ExcelError')
            new_row_data[ALL_EXPECTED_HEADERS.index("Age")] = profile_age if profile_age else data.get('active_age',
                                                                                                       'N/A_ExcelError')
            ws.append(new_row_data)
            row_to_update = ws.max_row

        ws.cell(row=row_to_update, column=ALL_EXPECTED_HEADERS.index("Stroop Part1 Time (s)") + 1).value = p1_time
        ws.cell(row=row_to_update, column=ALL_EXPECTED_HEADERS.index("Stroop Part1 Errors") + 1).value = p1_errors
        ws.cell(row=row_to_update, column=ALL_EXPECTED_HEADERS.index("Stroop Part2 Time (s)") + 1).value = p2_time
        ws.cell(row=row_to_update, column=ALL_EXPECTED_HEADERS.index("Stroop Part2 Errors") + 1).value = p2_errors
        ws.cell(row=row_to_update, column=ALL_EXPECTED_HEADERS.index("Stroop Part3 Time (s)") + 1).value = p3_time
        ws.cell(row=row_to_update, column=ALL_EXPECTED_HEADERS.index("Stroop Part3 Errors") + 1).value = p3_errors
        ws.cell(row=row_to_update,
                column=ALL_EXPECTED_HEADERS.index("Stroop - Interrupted") + 1).value = interruption_status_stroop

        wb.save(EXCEL_FILENAME)
        logger.info(f"Stroop results for UID {unique_id} saved/updated. Interrupted: {is_interrupted}")

        current_state_for_summary = await state.get_state()
        if current_state_for_summary is not None:
            summary_text_stroop = f"Результаты Теста Струпа {'<b>ПРЕРВАНЫ</b>' if is_interrupted else '<b>СОХРАНЕНЫ</b> (в разработке)'}."
            if is_interrupted and p1_time is None and p1_errors is None:
                summary_text_stroop = f"Тест Струпа <b>ПРЕРВАН</b> досрочно. Результаты не зафиксированы."
            await trigger_event_message.answer(summary_text_stroop, parse_mode=ParseMode.HTML)

    except Exception as e:
        logger.error(f"Error saving Stroop results to Excel for UID {unique_id}: {e}")
        current_state_for_error_msg = await state.get_state()
        if current_state_for_error_msg is not None:
            await trigger_event_message.answer("Произошла ошибка при сохранении результатов Теста Струпа.")


async def cleanup_stroop_ui(state: FSMContext, bot_instance: Bot,
                            final_text: str = "Тест Струпа завершен или отменен."):
    data = await state.get_data()
    chat_id = data.get('stroop_chat_id')
    main_message_id = data.get('stroop_main_message_id')

    if chat_id and main_message_id:
        try:
            await bot_instance.edit_message_text(
                text=final_text, chat_id=chat_id, message_id=main_message_id, reply_markup=None
            )
        except TelegramBadRequest:
            pass

    stroop_operational_keys_for_fsm_cleanup = [
        'stroop_chat_id', 'stroop_main_message_id', 'stroop_current_trial',
        'stroop_part1_time_total', 'stroop_part1_errors_total',
        'stroop_part2_time_total', 'stroop_part2_errors_total',
        'stroop_part3_time_total', 'stroop_part3_errors_total',
        'stroop_stimuli_colors'
    ]
    current_fsm_data = await state.get_data()
    data_after_stroop_message_cleanup = {k: v for k, v in current_fsm_data.items() if
                                         k not in stroop_operational_keys_for_fsm_cleanup}
    await state.set_data(data_after_stroop_message_cleanup)


# --- Test Registry ---
async def check_if_corsi_results_exist(profile_unique_id: int) -> bool:
    if not profile_unique_id: return False
    try:
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        corsi_cols_indices = [
            ALL_EXPECTED_HEADERS.index("Corsi - Max Correct Sequence Length") + 1,
            ALL_EXPECTED_HEADERS.index("Corsi - Avg Time Per Element (s)") + 1,
            ALL_EXPECTED_HEADERS.index("Corsi - Sequence Times Detail") + 1
        ]
        for row_cells_tuple in ws.iter_rows(min_row=2):
            if row_cells_tuple[1].value == profile_unique_id:
                for col_idx in corsi_cols_indices:
                    if ws.cell(row=row_cells_tuple[0].row, column=col_idx).value is not None:
                        return True
        return False
    except Exception as e:
        logger.error(f"Excel check error for Corsi results (UID {profile_unique_id}): {e}")
        return False


async def check_if_stroop_results_exist(profile_unique_id: int) -> bool:
    if not profile_unique_id: return False
    try:
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        stroop_cols_indices = [
            ALL_EXPECTED_HEADERS.index("Stroop Part1 Time (s)") + 1,
            ALL_EXPECTED_HEADERS.index("Stroop Part1 Errors") + 1,
        ]
        for row_cells_tuple in ws.iter_rows(min_row=2):
            if row_cells_tuple[1].value == profile_unique_id:
                for col_idx in stroop_cols_indices:
                    if ws.cell(row=row_cells_tuple[0].row, column=col_idx).value is not None:
                        return True
        return False
    except Exception as e:
        logger.error(f"Excel check error for Stroop results (UID {profile_unique_id}): {e}")
        return False


TEST_REGISTRY = {
    "initiate_corsi_test": {
        "name": "Тест Корси",
        "fsm_group_class": CorsiTestStates,
        "start_function": start_corsi_test,
        "save_function": save_corsi_results,
        "cleanup_function": cleanup_corsi_messages,
        "results_exist_check": check_if_corsi_results_exist,
        "requires_active_profile": True,
    },
    "initiate_stroop_test": {
        "name": "Тест Струпа (в разработке)",
        "fsm_group_class": StroopTestStates,
        "start_function": start_stroop_test,
        "save_function": save_stroop_results,
        "cleanup_function": cleanup_stroop_ui,
        "results_exist_check": check_if_stroop_results_exist,
        "requires_active_profile": True,
    }
}


# --- /stoptest Command Handler ---
@dp.message(Command("stoptest"))
async def stop_test_command_handler(message: Message, state: FSMContext, called_from_test_button: bool = False):
    current_fsm_state_str = await state.get_state()
    active_test_key = None
    active_test_config = None

    if current_fsm_state_str:
        for test_key, config in TEST_REGISTRY.items():
            if current_fsm_state_str.startswith(config["fsm_group_class"].__name__):
                active_test_key = test_key
                active_test_config = config
                break

    if active_test_config:
        if not called_from_test_button:
            await message.answer(f"Останавливаю тест: {active_test_config['name']}...")

        await active_test_config["save_function"](message, state, is_interrupted=True)
        await active_test_config["cleanup_function"](state, bot,
                                                     final_text=f"Тест {active_test_config['name']} был прерван.")

        fsm_data_after_test_cleanup = await state.get_data()
        main_profile_data_to_keep = {}
        if "active_unique_id" in fsm_data_after_test_cleanup:
            main_profile_data_to_keep["active_unique_id"] = fsm_data_after_test_cleanup.get("active_unique_id")
            main_profile_data_to_keep["active_name"] = fsm_data_after_test_cleanup.get("active_name")
            main_profile_data_to_keep["active_age"] = fsm_data_after_test_cleanup.get("active_age")
            main_profile_data_to_keep["active_telegram_id"] = fsm_data_after_test_cleanup.get("active_telegram_id")

        await state.set_state(None)

        if main_profile_data_to_keep.get("active_unique_id"):
            await state.set_data(main_profile_data_to_keep)
            await send_main_action_menu(message, ACTION_SELECTION_KEYBOARD_RETURNING, state=state)
            logger.info(
                f"Test '{active_test_config['name']}' stopped. User {message.from_user.id} (UID: {main_profile_data_to_keep.get('active_unique_id')}) returned to menu.")
        else:
            await message.answer("Тест остановлен. Ваш профиль не активен, пожалуйста, используйте /start.")
            logger.warning(
                f"Test '{active_test_config['name']}' stopped, but no active_profile data found to restore after cleanup. User {message.from_user.id}")
            await state.clear()
    elif not called_from_test_button:
        await message.answer("Нет активного теста для остановки. Вы можете выбрать тест из меню (команда /start).")


# --- Test Initiation and Overwrite Confirmation ---
@dp.callback_query(F.data == "select_specific_test")
async def on_select_specific_test_callback(cb: CallbackQuery, state: FSMContext):
    active_profile = await get_active_profile_from_fsm(state)
    if not active_profile:
        await cb.answer("Ваш профиль не активен. Пожалуйста, пройдите /start.", show_alert=True)
        try:
            await cb.message.edit_reply_markup(reply_markup=None)
        except TelegramBadRequest:
            pass
        return

    buttons = []
    for test_key, config in TEST_REGISTRY.items():
        if config.get("requires_active_profile", True):
            buttons.append([IKB(text=config["name"], callback_data=f"select_test_{test_key}")])

    if not buttons:
        await cb.message.edit_text("Нет доступных тестов для выбора.", reply_markup=None)
        await cb.answer()
        return

    await cb.answer()
    kbd = InlineKeyboardMarkup(inline_keyboard=buttons)
    try:
        await cb.message.edit_text("Выберите тест:", reply_markup=kbd)
    except TelegramBadRequest:
        await cb.message.answer("Выберите тест:", reply_markup=kbd)


@dp.callback_query(F.data.startswith("select_test_"))
async def on_test_selected_callback(cb: CallbackQuery, state: FSMContext):
    test_key_selected = cb.data.replace("select_test_", "")

    if test_key_selected not in TEST_REGISTRY:
        await cb.answer("Выбранный тест не найден.", show_alert=True)
        logger.warning(f"Unknown test key selected: {test_key_selected}")
        return

    test_config = TEST_REGISTRY[test_key_selected]
    active_profile = await get_active_profile_from_fsm(state)

    if not active_profile and test_config.get("requires_active_profile"):
        await cb.answer("Для этого теста требуется активный профиль. Пожалуйста, /start.", show_alert=True)
        return

    await cb.answer()
    await state.update_data(pending_test_key_for_overwrite=test_key_selected)

    results_exist = await test_config["results_exist_check"](
        active_profile.get("unique_id")) if active_profile else False

    if results_exist:
        confirm_kbd = InlineKeyboardMarkup(inline_keyboard=[
            [IKB(text="Да, перезаписать", callback_data="confirm_overwrite_test_results")],
            [IKB(text="Нет, отмена", callback_data="cancel_overwrite_test_results")]
        ])
        try:
            msg = await cb.message.edit_text(
                f"У вас есть сохраненные результаты для теста '{test_config['name']}'. Перезаписать их?",
                reply_markup=confirm_kbd
            )
            await state.update_data(overwrite_confirmation_message_id=msg.message_id)
        except TelegramBadRequest:
            msg = await cb.message.answer(
                f"У вас есть сохраненные результаты для теста '{test_config['name']}'. Перезаписать их?",
                reply_markup=confirm_kbd
            )
            await state.update_data(overwrite_confirmation_message_id=msg.message_id)
        await state.set_state(UserData.waiting_for_test_overwrite_confirmation)
    else:
        message_to_edit_or_send_new = cb.message
        if message_to_edit_or_send_new:
            try:
                await message_to_edit_or_send_new.edit_text(f"Подготовка к тесту: {test_config['name']}...",
                                                            reply_markup=None)
            except TelegramBadRequest:
                pass
        await test_config["start_function"](cb, state, active_profile)


@dp.callback_query(F.data == "confirm_overwrite_test_results", UserData.waiting_for_test_overwrite_confirmation)
async def handle_confirm_overwrite_test_results(cb: CallbackQuery, state: FSMContext):
    fsm_data = await state.get_data()
    test_key_to_start = fsm_data.get("pending_test_key_for_overwrite")

    if not test_key_to_start or test_key_to_start not in TEST_REGISTRY:
        await cb.answer("Ошибка: не удалось определить тест.", show_alert=True)
        await state.set_state(None)
        if cb.message:
            try:
                await cb.message.delete()
            except:
                pass
        return

    await cb.answer()
    test_config = TEST_REGISTRY[test_key_to_start]
    active_profile = await get_active_profile_from_fsm(state)

    if not active_profile and test_config.get("requires_active_profile"):
        await cb.message.answer("Ошибка: ваш профиль не активен. Пожалуйста, /start.")
        await state.set_state(None)
        return

    if cb.message:
        try:
            await cb.message.edit_text(f"Подготовка к тесту: {test_config['name']} (перезапись)...", reply_markup=None)
        except TelegramBadRequest:
            pass
    await state.update_data(overwrite_confirmation_message_id=None, pending_test_key_for_overwrite=None)
    await test_config["start_function"](cb, state, active_profile)


@dp.callback_query(F.data == "cancel_overwrite_test_results", UserData.waiting_for_test_overwrite_confirmation)
async def handle_cancel_overwrite_test_results(cb: CallbackQuery, state: FSMContext):
    await cb.answer("Отменено.", show_alert=False)
    fsm_data = await state.get_data()
    test_key_pending = fsm_data.get("pending_test_key_for_overwrite")
    test_name = TEST_REGISTRY[test_key_pending][
        'name'] if test_key_pending and test_key_pending in TEST_REGISTRY else "теста"

    try:
        await cb.message.edit_text(f"Запуск {test_name} отменен.", reply_markup=None)
    except TelegramBadRequest:
        pass

    await state.update_data(overwrite_confirmation_message_id=None, pending_test_key_for_overwrite=None)
    await state.set_state(None)
    await send_main_action_menu(cb.message, ACTION_SELECTION_KEYBOARD_RETURNING, state=state)


@dp.callback_query(F.data == "run_test_battery")
async def on_run_test_battery_callback(cb: CallbackQuery, state: FSMContext):
    await cb.answer("Функция 'Батарея тестов' в разработке.", show_alert=True)


# --- Registration and Main Menu Handlers ---
@dp.message(CommandStart())
async def start_command_handler(message: Message, state: FSMContext):
    await state.clear()
    await state.set_state(UserData.waiting_for_first_time_response)
    first_time_kbd = InlineKeyboardMarkup(
        inline_keyboard=[
            [IKB(text="Да (зарегистрироваться)", callback_data="user_is_new")],
            [IKB(text="Нет (войти по UID)", callback_data="user_is_returning")],
        ]
    )
    await message.answer("Вы впервые пользуетесь ботом?", reply_markup=first_time_kbd)


@dp.callback_query(F.data == "user_is_new", UserData.waiting_for_first_time_response)
async def handle_user_is_new_callback(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    try:
        await callback.message.edit_reply_markup(reply_markup=None)
    except TelegramBadRequest:
        pass
    await state.set_state(UserData.waiting_for_name)
    await callback.message.answer('Привет! Давайте начнем. Как вас зовут?')


@dp.callback_query(F.data == "user_is_returning", UserData.waiting_for_first_time_response)
async def handle_user_is_returning_callback(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    try:
        await callback.message.edit_reply_markup(reply_markup=None)
    except TelegramBadRequest:
        pass
    await state.set_state(UserData.waiting_for_unique_id)
    await callback.message.answer("Введите ваш уникальный идентификатор (UID):")


@dp.message(UserData.waiting_for_unique_id)
async def process_unique_id_input(message: Message, state: FSMContext):
    try:
        entered_unique_id_str = message.text.strip()
        if not entered_unique_id_str.isdigit():
            await message.answer("UID должен быть числом. Пожалуйста, введите корректный UID.")
            return
        entered_unique_id = int(entered_unique_id_str)
    except ValueError:
        await message.answer("Пожалуйста, введите корректный числовой UID.")
        return

    try:
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        user_profile_data = None
        for row_idx, row_cells_tuple in enumerate(ws.iter_rows(min_row=2), start=2):
            if row_cells_tuple[1].value == entered_unique_id:
                user_profile_data = {
                    "active_unique_id": entered_unique_id,
                    "active_telegram_id": row_cells_tuple[0].value,
                    "active_name": str(row_cells_tuple[2].value),
                    "active_age": str(row_cells_tuple[3].value),
                }
                logger.info(f"User authenticated via UID: {entered_unique_id}. Profile: {user_profile_data}")
                break

        if user_profile_data:
            await state.set_data(user_profile_data)
            await state.set_state(None)
            await message.answer(f"Авторизация прошла успешно, {user_profile_data['active_name']}!")
            await send_main_action_menu(message, ACTION_SELECTION_KEYBOARD_RETURNING, state=state)
        else:
            kbd = InlineKeyboardMarkup(inline_keyboard=[
                [IKB(text="Попробовать снова", callback_data="try_id_again")],
                [IKB(text="Зарегистрироваться как новый", callback_data="register_new_after_fail")]
            ])
            await message.answer("Уникальный идентификатор (UID) не найден.", reply_markup=kbd)
    except Exception as e:
        logger.error(f"Error during UID check for '{message.text}': {e}")
        await message.answer("Произошла ошибка при проверке UID. Попробуйте позже или свяжитесь с администратором.")


@dp.callback_query(F.data == "try_id_again", UserData.waiting_for_unique_id)
async def handle_try_id_again_callback(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    try:
        await callback.message.edit_reply_markup(reply_markup=None)
    except TelegramBadRequest:
        pass
    await callback.message.answer("Введите ваш уникальный идентификатор (UID):")


@dp.callback_query(F.data == "register_new_after_fail", UserData.waiting_for_unique_id)
async def handle_register_new_after_fail_callback(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    try:
        await callback.message.edit_reply_markup(reply_markup=None)
    except TelegramBadRequest:
        pass
    await state.set_state(UserData.waiting_for_name)
    await callback.message.answer('Как вас зовут?')


@dp.message(UserData.waiting_for_name)
async def process_name_input(message: Message, state: FSMContext):
    name_input = message.text.strip()
    if not name_input:
        await message.answer("Имя не может быть пустым. Пожалуйста, введите ваше имя.")
        return
    await state.update_data(name_for_registration=name_input)
    await state.set_state(UserData.waiting_for_age)
    await message.answer('Отлично! Теперь введите ваш возраст (цифрами).')


@dp.message(UserData.waiting_for_age)
async def process_age_input(message: Message, state: FSMContext):
    age_input = message.text.strip()
    if not age_input.isdigit() or not (0 < int(age_input) < 120):
        await message.answer("Пожалуйста, введите корректный возраст цифрами (например, 25).")
        return

    fsm_data = await state.get_data()
    name_to_register = fsm_data.get('name_for_registration')
    age_to_register = int(age_input)
    current_telegram_id = message.from_user.id

    new_unique_id = None
    try:
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        existing_ids = {row[1].value for row in ws.iter_rows(min_row=2) if row[1].value is not None}

        min_uid, max_uid = 1000000, 9999999
        if len(existing_ids) >= (max_uid - min_uid + 1):
            await message.answer(
                "Критическая ошибка: не удалось сгенерировать UID, все идентификаторы исчерпаны. Свяжитесь с администратором.")
            logger.critical("All 7-digit UIDs seem to be exhausted.")
            await state.clear()
            return

        attempts = 0
        while attempts < 1000:
            new_unique_id = random.randint(min_uid, max_uid)
            if new_unique_id not in existing_ids:
                break
            attempts += 1
        if new_unique_id is None or new_unique_id in existing_ids:
            await message.answer(
                "Критическая ошибка: не удалось сгенерировать уникальный UID после множества попыток. Свяжитесь с администратором.")
            logger.critical(f"Failed to generate a unique 7-digit UID after {attempts} attempts.")
            await state.clear()
            return

        new_user_row = [''] * len(ALL_EXPECTED_HEADERS)
        new_user_row[ALL_EXPECTED_HEADERS.index("Telegram ID")] = current_telegram_id
        new_user_row[ALL_EXPECTED_HEADERS.index("Unique ID")] = new_unique_id
        new_user_row[ALL_EXPECTED_HEADERS.index("Name")] = name_to_register
        new_user_row[ALL_EXPECTED_HEADERS.index("Age")] = age_to_register

        ws.append(new_user_row)
        wb.save(EXCEL_FILENAME)
        logger.info(
            f"New user registered: TG ID: {current_telegram_id}, UID: {new_unique_id}, Name: {name_to_register}, Age: {age_to_register}")

        active_profile_data = {
            'active_telegram_id': current_telegram_id,
            'active_unique_id': new_unique_id,
            'active_name': name_to_register,
            'active_age': age_to_register,
        }
        await state.set_data(active_profile_data)
        await state.set_state(None)

        await message.answer(
            f"Спасибо, {name_to_register}! Вы успешно зарегистрированы.\n"
            f"<b>Ваш Уникальный Идентификатор (UID): {new_unique_id}</b>\n"
            f"Запомните или запишите его, он понадобится для входа в следующий раз."
        )
        await send_main_action_menu(message, ACTION_SELECTION_KEYBOARD_NEW, state=state)

    except Exception as e:
        logger.error(f"Error during new user registration (UID generation or Excel save): {e}", exc_info=True)
        await message.answer(
            "Произошла ошибка во время регистрации. Пожалуйста, попробуйте позже или свяжитесь с администратором.")
        await state.clear()


# --- Utility Handlers ---
@dp.message(Command("mydata"))
async def show_my_data_command(message: Message, state: FSMContext):
    active_profile = await get_active_profile_from_fsm(state)
    if not active_profile:
        await message.answer("Ваш профиль не активен. Пожалуйста, войдите или зарегистрируйтесь через /start.")
        return

    uid_to_show = active_profile.get("unique_id")
    response_lines = [f"Данные для активного профиля UID: <b>{uid_to_show}</b>"]

    try:
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        profile_found_in_excel = False
        for row_cells_tuple in ws.iter_rows(min_row=2):
            if row_cells_tuple[1].value == uid_to_show:
                profile_found_in_excel = True
                for i, header_name in enumerate(ALL_EXPECTED_HEADERS):
                    cell_value = row_cells_tuple[i].value
                    if "Interrupted" in header_name and cell_value is not None:
                        display_value = "Да" if cell_value == "Да" else ("Нет" if cell_value == "Нет" else cell_value)
                    else:
                        display_value = cell_value if cell_value is not None else "нет данных"
                    response_lines.append(f"<b>{header_name}:</b> {display_value}")
                break
        if not profile_found_in_excel:
            response_lines.append("Профиль с таким UID не найден в базе данных (Excel). Это неожиданно.")
            logger.warning(f"/mydata: Active UID {uid_to_show} from FSM not found in Excel.")

    except FileNotFoundError:
        response_lines.append("Файл данных не найден. Свяжитесь с администратором.")
        logger.error(f"/mydata: Excel file '{EXCEL_FILENAME}' not found.")
    except Exception as e:
        response_lines.append("Ошибка при загрузке данных. Свяжитесь с администратором.")
        logger.error(f"Error loading Excel for /mydata (UID: {uid_to_show}): {e}")

    await message.answer("\n".join(response_lines), parse_mode=ParseMode.HTML)


@dp.message(Command("export"))
async def export_data_to_excel_command(message: Message, state: FSMContext):
    if os.path.exists(EXCEL_FILENAME):
        try:
            await message.reply_document(FSInputFile(EXCEL_FILENAME), caption="Данные пользователей.")
        except Exception as e:
            logger.error(f"Error sending Excel file: {e}")
            await message.answer("Не удалось отправить файл. Попробуйте позже.")
    else:
        await message.answer(f"Файл данных '{EXCEL_FILENAME}' не найден.")


@dp.message(Command("restart"))
async def command_restart_bot_session_handler(message: Message, state: FSMContext):
    current_fsm_state_str = await state.get_state()
    active_test_key = None
    if current_fsm_state_str:
        for test_key, config in TEST_REGISTRY.items():
            if current_fsm_state_str.startswith(config["fsm_group_class"].__name__):
                active_test_key = test_key
                break
    if active_test_key and TEST_REGISTRY[active_test_key].get("cleanup_function"):
        await TEST_REGISTRY[active_test_key]["cleanup_function"](state, bot,
                                                                 final_text=f"Тест был остановлен командой /restart.")

    await state.clear()
    await message.answer(
        "Все текущие операции остановлены и ваш профиль сброшен.\n"
        "Используйте /start, чтобы начать заново (войти или зарегистрироваться)."
    )


@dp.callback_query(F.data == "logout_profile")
async def logout_profile_callback(cb: CallbackQuery, state: FSMContext):
    await cb.answer("Профиль сброшен.", show_alert=True)
    await state.clear()
    try:
        await cb.message.edit_text(
            "Ваш текущий профиль был сброшен. Вы можете войти снова или зарегистрироваться через /start.",
            reply_markup=None
        )
    except TelegramBadRequest:
        await cb.message.answer(
            "Ваш текущий профиль был сброшен. Вы можете войти снова или зарегистрироваться через /start."
        )

    await state.set_state(UserData.waiting_for_first_time_response)
    first_time_kbd = InlineKeyboardMarkup(inline_keyboard=[
        [IKB(text="Да (зарегистрироваться)", callback_data="user_is_new")],
        [IKB(text="Нет (войти по UID)", callback_data="user_is_returning")],
    ])
    await cb.message.answer("Вы впервые пользуетесь ботом?", reply_markup=first_time_kbd)


# --- Main Bot Execution ---
async def main():
    initialize_excel_file()
    logger.info("Bot starting...")

    dp.callback_query.register(handle_corsi_button_press, F.data.startswith("corsi_button_"),
                               CorsiTestStates.waiting_for_user_sequence)
    dp.callback_query.register(on_corsi_restart_current_test, F.data == "corsi_stop_this_attempt",
                               StateFilter(CorsiTestStates))

    dp.callback_query.register(handle_stroop_part1_response, F.data == "stroop_p1_next", StroopTestStates.part1_display)

    await bot.delete_webhook(drop_pending_updates=True)
    await dp.start_polling(bot)


if __name__ == '__main__':
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logging.info('Bot stopped by user.')
    except Exception as e:
        logging.error(f"Unhandled main exception: {e}", exc_info=True)
